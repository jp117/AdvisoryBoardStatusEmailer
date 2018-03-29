import openpyxl
import os
import datetime

#When I move the spreadsheet to the server, change the working directory to where the spreadsheet is stored
#os.chdir('c:\\users\\John Paradise\\MyPythonScripts\\AdvisoryBoard')

workbook = openpyxl.load_workbook('AdvisoryBoardTracking.xlsx')
newSub = workbook['NewSubmissions']
pendingSub = workbook['PendingReSubmissions']
meetingWB = workbook['MeetingSchedule']

#Finds next row in sheet
def nextrow(sheet):
	i = 1
	row = sheet.cell(row=i, column=1).value
	while row != None:
		i += 1
		row = sheet.cell(row=i, column=1).value
	return i

#finds the next Advisory Board meeting
def nextmeeting():
	today = datetime.datetime.now().strftime('%m/%d/%Y')
	i = 1
	subDate = datetime.datetime.strftime(meetingWB.cell(row=i, column=2).value, '%m/%d/%Y')
	while today > subDate:
		i += 1
		subDate = datetime.datetime.strftime(meetingWB.cell(row=i, column=2).value, '%m/%d/%Y')
	return i

#Makes a list of lists of all jobs that are awaiting submission 
def newSubList(sheet):
	datalistoflist = []
	i = 2
	for i in range (i, nextrow(sheet)):
		data = sheet.cell(row=i, column=13).value
		if data == None or data.lower() != "yes":
			datalist = []
			for x in range (1,15):
				datalist.append(sheet.cell(row=i, column=x).value)
			datalistoflist.append(datalist)
	return datalistoflist

#Makes a list of lists of all jobs that are pending and awaiting resubmission
def pendingSubList(sheet):
	datalistoflist = []
	i = 2
	for i in range (i, nextrow(sheet)):
		data = sheet.cell(row=i, column=10).value
		if data == None or data.lower() != "yes":
			datalist = []
			for x in range (1,13):
				datalist.append(sheet.cell(row=i, column=x).value)
			datalistoflist.append(datalist)
	return datalistoflist

#Assembles the email body for submission from NewSubmission worksheet
def EmailNewSubBody():
	i = len(newSubList(newSub))
	subbody = """\
<html>
	<head>
	</head>
		<body>
<p><span id="01"><b>This is an update of our new submissions to the advisory board</b>.</span><br />
The next meeting we can submit to is <b>"""
	subbody += str(datetime.datetime.strftime(meetingWB.cell(row=nextmeeting(), column=1).value, '%m/%d/%Y')) + "</b>, which means we must submit on <b>" + str(datetime.datetime.strftime(meetingWB.cell(row=nextmeeting(), column=2).value, '%m/%d/%Y'))
	subbody += """</b><br />
If you have a job that is not listed that needs to be submitted, please email John or Gina so that it can be added to the list.<br />
If there is a problem with the target submission date, please let Gina or John know so the schedule can be adjusted.<br />
If a check and letter is needed, please see engineering for the drawing # and request the check and letter for your job.<br /><br /></p>
""" 
	for x in range (0, i):
		subbody += "<b>SO#: " + str(newSubList(newSub)[x][0]) + "</b><br />"
		subbody += "Contractor: " + str(newSubList(newSub)[x][1]) + "<br />"
		if newSubList(newSub)[x][2] == None:
			subbody += "Job: " + str(newSubList(newSub)[x][3]) + "<br />"
		elif newSubList(newSub)[x][3] == None:
			subbody += "Job: " + str(newSubList(newSub)[x][2]) + '<br />'
		else:
			subbody += "Job: " + str(newSubList(newSub)[x][2]) + " at " + str(newSubList(newSub)[x][3]) + "<br />"
		if newSubList(newSub)[x][4] == None:
			subbody += "Sub#: REQUEST SUBMISSION NUMBER" + "<br />"
		else:
			subbody += "Sub#: " + str(newSubList(newSub)[x][4]) + "<br />"
		if newSubList(newSub)[x][5] == None or newSubList(newSub)[x][5].lower() == "no":
			subbody += "Check and Letter: REQUEST CHECK AND LETTER" + "<br />"
		elif newSubList(newSub)[x][5].lower() == "yes":
			subbody += "Check and Letter: Recieved<br />"
		else:
			subbody += "Check and Letter: " + str(newSubList(newSub)[x][5]) + "<br />"
		if newSubList(newSub)[x][7] != None:
			subbody += "Target Meeting: " + str(newSubList(newSub)[x][7]) + "<br />"
		else:
			subbody += "No target meeting set <br />"
		subbody += "Salesman: " + str(newSubList(newSub)[x][9]) + "<br />"
		subbody += "Engineer: " + str(newSubList(newSub)[x][8]) + '<br />'
		if newSubList(newSub)[x][12] == None or newSubList(newSub)[x][12].lower() != "yes":
			subbody += "Not yet ready to submit<br />"
		else: 
			subbody =+ "READY TO SUBMIT<br />"
		subbody += "<br />"
	return subbody

#Take the email body from EmailNewSubBody and adds the data from Pending part of the workbook
def EmailPendingSubBody():
	i = len(pendingSubList(pendingSub))
	subbody = EmailNewSubBody()
	subbody += "<b>PENDING SUBMISSIONS BELOW</b><br /><br />"
	for x in range (0,i):
		subbody += "<b>SO#: " + str(pendingSubList(pendingSub)[x][0]) + "</b><br />"
		subbody += "Contractor: " + str(pendingSubList(pendingSub)[x][1]) + "<br />"
		if pendingSubList(pendingSub)[x][2] == None:
			subbody += "Job " + str(pendingSubList(pendingSub)[x][3]) + '<br />'
		elif pendingSubList(pendingSub)[x][3] == None:
			subbody += "Job " + str(pendingSubList(pendingSub)[x][2]) + '<br />'
		else:
			subbody += "Job " + str(pendingSubList(pendingSub)[x][2]) + " at " + str(pendingSubList(pendingSub)[x][3]) + "<br />"
		subbody += "Sub #: " + str(pendingSubList(pendingSub)[x][4]) + '<br />'
		subbody += "Advisory Board Meeting Submitted: " + str(pendingSubList(pendingSub)[x][11]) + "<br />"
		subbody += "Salesman: " + str(pendingSubList(pendingSub)[x][6]) + '<br />'
		subbody += "Engineer: " + str(pendingSubList(pendingSub)[x][5]) + '<br />'
		if pendingSubList(pendingSub)[x][8] == None or pendingSubList(pendingSub)[x][8].lower() != "yes":
			subbody += "Not ready to resubmit yet<br />"
		else:
			subbody += "READY TO RESUBMIT<br />"
		subbody += "<br /></body></html>"
	return subbody
